import tkinter as tk
import threading
import serial
import time
import datetime
import queue
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
import sys

class SerialLogic:
    def __init__(self, serial_frame, command_frame, monitor_frame):
        # Store references to UI frames
        self.serial_frame = serial_frame
        self.command_frame = command_frame
        self.monitor_frame = monitor_frame
        
        # Serial connection settings
        self.serial_conn = None
        self.is_connected = False
        self.port = None
        self.baudrate = None
        self.available_ports = []
        
        # Command execution settings
        self.is_running = False
        self.should_stop = False
        self.command_queue = queue.Queue()
        self.current_cycle = 0
        self.total_cycles = 0
        self.start_time = None
        self.elapsed_time_thread = None
        self.command_thread = None

        
        # Result tracking
        self.results_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "RESULTS.xlsx")
        self.setup_excel_file()
        
        # Connect UI elements to logic
        self.setup_ui_connections()
        self.scan_ports()
        self.update_ui_state()
        
    def clear_everything(self):
        """Clear everything and reset the application state"""
        # Clear the monitor
        self.monitor_frame.clearMonitor()
        
        # Disconnect if connected
        if self.is_connected:
            self.disconnect_serial()
        
        # Clear the command table
        for item in self.command_frame.commandTable.get_children():
            self.command_frame.commandTable.delete(item)
        
        # Reset cycle count and elapsed time
        self.command_frame.cycleVar.set("1")
        self.command_frame.elapsedTimeVar.set("00:00:00")
        
        # Clear command entry
        self.command_frame.commandVar.set("")     

        for item_id in self.command_frame.commandTable.get_children():
            values = list(self.command_frame.commandTable.item(item_id)['values'])
            values[2] = ""  # Clear the response column
            self.command_frame.commandTable.item(item_id, values=values)
    
    def setup_ui_connections(self):
        """Connect UI elements to their respective functions"""
        # Serial frame connections
        self.serial_frame.refreshBtn.configure(command=self.scan_ports)
        self.serial_frame.connectBtn.configure(command=self.toggle_connection)
        
        # Command frame connections
        self.command_frame.runStopBtn.configure(command=self.toggle_run_stop)

        self.monitor_frame.clearBtn.configure(command=self.clear_everything)
        
        # Initialize UI state
        self.disable_command_controls()
    
    def scan_ports(self):
        """Scan for available COM ports"""
        try:
            import serial.tools.list_ports
            self.available_ports = [port.device for port in serial.tools.list_ports.comports()]
            # Update the combobox values
            self.serial_frame.comPortCombo['values'] = self.available_ports
            
            if self.available_ports:
                self.serial_frame.comPortCombo.current(0)
            else:
                self.serial_frame.comPortVar.set("")
                
            self.monitor_frame.appendToMonitor(f"Found {len(self.available_ports)} ports: {', '.join(self.available_ports)}", "SYS")
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"Error scanning ports: {str(e)}", "ERR")
    
    def toggle_connection(self):
        """Toggle the serial connection state"""
        if not self.is_connected:
            self.connect_serial()
        else:
            self.disconnect_serial()
    
    def connect_serial(self):
        """Connect to the selected serial port"""
        try:
            self.port = self.serial_frame.comPortVar.get()
            self.baudrate = int(self.serial_frame.baudRateVar.get())
            
            if not self.port:
                self.monitor_frame.appendToMonitor("No COM port selected", "ERR")
                return
            
            self.serial_conn = serial.Serial(
                port=self.port,
                baudrate=self.baudrate,
                timeout=1
            )
            
            self.is_connected = True
            self.serial_frame.connectVar.set("Disconnect")
            self.monitor_frame.appendToMonitor(f"Connected to {self.port} at {self.baudrate} baud", "SYS")
            
            # Enable command controls
            self.enable_command_controls()
            
            # Start the serial monitoring thread
            threading.Thread(target=self.serial_monitor_thread, daemon=True).start()
            
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"Connection error: {str(e)}", "ERR")
    
    def disconnect_serial(self):
        """Disconnect from the serial port"""
        try:
            if self.serial_conn and self.serial_conn.is_open:
                # Stop any running operations
                if self.is_running:
                    self.toggle_run_stop()
                
                self.serial_conn.close()
            
            self.is_connected = False
            self.serial_frame.connectVar.set("Connect")
            self.monitor_frame.appendToMonitor(f"Disconnected from {self.port}", "SYS")
            
            # Disable command controls
            self.disable_command_controls()
            
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"Disconnection error: {str(e)}", "ERR")
    
    def toggle_run_stop(self):
        """Toggle between Run and Stop states"""
        if not self.is_running:
            self.start_command_execution()
        else:
            self.stop_command_execution()
    
    def start_command_execution(self):
        
        for item_id in self.command_frame.commandTable.get_children():
            values = list(self.command_frame.commandTable.item(item_id)['values'])
            values[2] = ""  # Clear the response column
            self.command_frame.commandTable.item(item_id, values=values)
        """Start executing commands from the table"""
        # Check if there are commands in the table
        commands = self.get_commands_from_table()
        if not commands:
            self.monitor_frame.appendToMonitor("No commands to execute", "SYS")
            return
        
        # Check if cycle count is valid
        try:
            cycles = int(self.command_frame.cycleVar.get())
            if cycles <= 0:
                self.monitor_frame.appendToMonitor("Cycle count must be greater than 0", "SYS")
                return
        except ValueError:
            self.monitor_frame.appendToMonitor("Invalid cycle count", "ERR")
            return
        
        # Set up execution variables
        self.is_running = True
        self.should_stop = False
        self.current_cycle = 0
        self.total_cycles = cycles
        self.command_frame.cycleProgressVar.set(f"0/{cycles}")
        
        # Start elapsed time counter
        self.start_time = time.time()
        self.elapsed_time_thread = threading.Thread(target=self.update_elapsed_time, daemon=True)
        self.elapsed_time_thread.start()
        
        # Update UI state
        self.command_frame.runStopBtn.configure(text="STOP", bootstyle="danger")
        self.disable_command_editing()
        
        # Start command execution thread
        self.command_thread = threading.Thread(target=self.execute_commands, args=(commands, cycles), daemon=True)
        self.command_thread.start()
    
    def stop_command_execution(self):
        """Stop the command execution"""
        if self.serial_conn and self.serial_conn.is_open:
            # Send the "HALT" command before stopping execution
            self.serial_conn.write(b"HALT\n")
            self.monitor_frame.appendToMonitor("Sending: HALT", "TX")

        self.should_stop = True
        self.is_running = False
        self.command_frame.runStopBtn.configure(text="RUN", bootstyle="success")
        self.enable_command_editing()
        self.monitor_frame.appendToMonitor("Command execution stopped", "SYS")
    
    def execute_commands(self, commands, cycles):
        """Execute the commands in the table for the specified number of cycles"""
        try:
            for cycle in range(1, cycles + 1):
                if self.should_stop:
                    break
                
                self.current_cycle = cycle
                self.command_frame.cycleProgressVar.set(f"{cycle}/{cycles}")
                
                for i, (item_id, command) in enumerate(commands):
                    if self.should_stop:
                        break
                    
                    # Update status to yellow (waiting)
                    self.update_command_status(item_id, "WAITING", "yellow")
                    
                    # Send command
                    self.monitor_frame.appendToMonitor(f"Sending: {command}", "TX")
                    self.serial_conn.write(f"{command}\n".encode())
                    
                    # Wait for response with timeout
                    response = self.wait_for_response(30)

                    if response == "TIMEOUT":
                        # Timeout occurred
                        self.update_command_status(item_id, "TIMEOUT", "red")
                        self.update_command_response(item_id, "No response received (timeout)")
                        self.should_stop = True
                        break
                    elif response == "HALT":
                         # Stop execution if "HALT" command received
                        self.update_command_status(item_id, "HALT", "red")
                        self.update_command_response(item_id, "STOP execution")
                        self.should_stop = True
                        break

                    # Process response
                    response_str = response.decode('utf-8', errors='replace').strip()
                    self.monitor_frame.appendToMonitor(f"Received: {response_str}", "RX")
                    
                    # Update response in table
                    self.update_command_response(item_id, response_str)
                    
                    # Check for specific responses
                    if "_ERR" in response_str:
                        self.update_command_status(item_id, "ERROR", "red")
                        self.log_result(command, "ERROR", response_str)
                        self.should_stop = True
                        break
                    elif "_RDY" in response_str or "_REP" in response_str:
                        self.update_command_status(item_id, "SUCCESS", "green")
                        self.log_result(command, "SUCCESS", response_str)
                    else:
                        self.update_command_status(item_id, "UNKNOWN", "yellow")
                        self.log_result(command, "UNKNOWN", response_str)
            
            # Execution completed or stopped
            if not self.should_stop:
                self.monitor_frame.appendToMonitor("All commands executed successfully", "SYS")
            
            # Reset UI state
            self.is_running = False
            self.command_frame.runStopBtn.configure(text="RUN", bootstyle="success")
            self.enable_command_editing()
            
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"Execution error: {str(e)}", "ERR")
            self.is_running = False
            self.command_frame.runStopBtn.configure(text="RUN", bootstyle="success")
            self.enable_command_editing()
    
    def wait_for_response(self, timeout):
        """Wait for a response from the serial device with timeout"""
        start_time = time.time()
        response = b""
        
        while (time.time() - start_time) < timeout:
            if self.should_stop:
                return "HALT"
            
            if self.serial_conn.in_waiting > 0:
                new_data = self.serial_conn.read(self.serial_conn.in_waiting)
                response += new_data
                
                # Check if we have a complete response
                if b"_RDY" in response or b"_ERR" in response or b"_REP" in response:
                    return response
            
            # Small delay to prevent hogging CPU
            time.sleep(0.1)
        
        # Timeout occurred
        return "TIMEOUT"
    
    def serial_monitor_thread(self):
        """Thread to monitor incoming serial data when not executing commands"""
        try:
            while self.is_connected:
                # Only process incoming data when not executing commands
                if not self.is_running and self.serial_conn and self.serial_conn.is_open:
                    if self.serial_conn.in_waiting > 0:
                        data = self.serial_conn.read(self.serial_conn.in_waiting)
                        text = data.decode('utf-8', errors='replace')
                        self.monitor_frame.appendToMonitor(text, "RX")
                
                # Small delay to prevent hogging CPU
                time.sleep(0.1)
        except Exception as e:
            if self.is_connected:  # Only show error if we're supposed to be connected
                self.monitor_frame.appendToMonitor(f"Monitor error: {str(e)}", "ERR")
    
    def update_elapsed_time(self):
        """Update the elapsed time display"""
        while self.is_running:
            if self.start_time:
                elapsed = time.time() - self.start_time
                hours, remainder = divmod(int(elapsed), 3600)
                minutes, seconds = divmod(remainder, 60)
                time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                self.command_frame.elapsedTimeVar.set(time_str)
            time.sleep(0.1)
    
    def get_commands_from_table(self):
        """Get all commands from the table with their IDs"""
        commands = []
        for item_id in self.command_frame.commandTable.get_children():
            values = self.command_frame.commandTable.item(item_id)['values']
            if values and values[0]:  # Check if command is not empty
                commands.append((item_id, values[0]))
        return commands
    
    def update_command_status(self, item_id, status, color):
        """Update the status of a command in the table"""
        try:
            values = list(self.command_frame.commandTable.item(item_id)['values'])
            values[1] = status
            self.command_frame.commandTable.item(item_id, values=values)
            
            # Set row color based on status
            if color == "green":
                self.command_frame.commandTable.tag_configure(item_id, background="#c0ffc0")
                self.command_frame.commandTable.item(item_id, tags=(item_id,))
            elif color == "yellow":
                self.command_frame.commandTable.tag_configure(item_id, background="#ffffc0")
                self.command_frame.commandTable.item(item_id, tags=(item_id,))
            elif color == "red":
                self.command_frame.commandTable.tag_configure(item_id, background="#ffc0c0")
                self.command_frame.commandTable.item(item_id, tags=(item_id,))
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"UI update error: {str(e)}", "ERR")
    
    def update_command_response(self, item_id, response):
        """Update the response of a command in the table"""
        try:
            values = list(self.command_frame.commandTable.item(item_id)['values'])
            values[2] = response
            self.command_frame.commandTable.item(item_id, values=values)
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"UI update error: {str(e)}", "ERR")
    
    def setup_excel_file(self):
        """Set up the Excel results file"""
        try:
            if not os.path.exists(self.results_file):
                # Create new workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Results"
                
                # Add headers
                headers = ["COMMAND", "RESPONSE", "TIME"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                
                wb.save(self.results_file)
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"Excel setup error: {str(e)}", "ERR")
    
    def log_result(self, command, status, response):
        """Log the command result to the Excel file"""
        try:
            # Load workbook
            wb = openpyxl.load_workbook(self.results_file)
            ws = wb.active
            
            # Get current timestamp
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Add new row
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1).value = command
            ws.cell(row=next_row, column=2).value = response
            ws.cell(row=next_row, column=3).value = timestamp
            
            # Color based on status
            if status == "SUCCESS":
                for col in range(1, 4):
                    ws.cell(row=next_row, column=col).fill = PatternFill(
                        start_color='C0FFC0', end_color='C0FFC0', fill_type='solid')
            elif status == "ERROR":
                for col in range(1, 4):
                    ws.cell(row=next_row, column=col).fill = PatternFill(
                        start_color='FFC0C0', end_color='FFC0C0', fill_type='solid')
            elif status == "UNKNOWN" or status == "TIMEOUT":
                for col in range(1, 4):
                    ws.cell(row=next_row, column=col).fill = PatternFill(
                        start_color='FFFFC0', end_color='FFFFC0', fill_type='solid')
            
            # Save workbook
            wb.save(self.results_file)
        except Exception as e:
            self.monitor_frame.appendToMonitor(f"Excel logging error: {str(e)}", "ERR")
    
    def enable_command_controls(self):
        """Enable command-related UI controls"""
        self.command_frame.commandEntry.configure(state="normal")
        self.command_frame.cycleEntry.configure(state="normal")
        self.command_frame.addBtn.configure(state="normal")
        self.command_frame.deleteBtn.configure(state="normal")
        self.command_frame.updateBtn.configure(state="normal")
        self.command_frame.runStopBtn.configure(state="normal")
    
    def disable_command_controls(self):
        """Disable command-related UI controls"""
        self.command_frame.commandEntry.configure(state="disabled")
        self.command_frame.cycleEntry.configure(state="disabled")
        self.command_frame.addBtn.configure(state="disabled")
        self.command_frame.deleteBtn.configure(state="disabled")
        self.command_frame.updateBtn.configure(state="disabled")
        self.command_frame.runStopBtn.configure(state="disabled")
    
    def enable_command_editing(self):
        """Enable command editing controls"""
        self.command_frame.commandEntry.configure(state="normal")
        self.command_frame.cycleEntry.configure(state="normal")
        self.command_frame.addBtn.configure(state="normal")
        self.command_frame.deleteBtn.configure(state="normal")
        self.command_frame.updateBtn.configure(state="normal")
    
    def disable_command_editing(self):
        """Disable command editing controls"""
        self.command_frame.commandEntry.configure(state="disabled")
        self.command_frame.cycleEntry.configure(state="disabled")
        self.command_frame.addBtn.configure(state="disabled")
        self.command_frame.deleteBtn.configure(state="disabled")
        self.command_frame.updateBtn.configure(state="disabled")
    
    def update_ui_state(self):
        """Update UI state based on current connection status"""
        if self.is_connected:
            self.serial_frame.connectVar.set("Disconnect")
            self.enable_command_controls()
        else:
            self.serial_frame.connectVar.set("Connect")
            self.disable_command_controls()